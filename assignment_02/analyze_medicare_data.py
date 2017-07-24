 #libraries
import requests
import os
import zipfile
import openpyxl
import sqlite3
import glob
#import getpass
import pandas as pd
import numpy as np
#importing the library to convert dataframe into rows for excel workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import gc
#declaring links to download the required files
download_url = url="https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"
k_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"
staging_dir_name = "staging"
states_focused = {}
#Method for downloading the csv files
#extracted files are stored in staging directory
def csvfiles_download():
    os.mkdir(staging_dir_name)
    zip_file_name = os.path.join(staging_dir_name, "test.zip")
    r = requests.get(url)
    zf = open(zip_file_name, "wb")
    zf.write(r.content)
    zf.close()
    z = zipfile.ZipFile(zip_file_name, 'r')
    z.extractall(staging_dir_name)
    z.close()
    xf = open("hospital_ranking_focus_states.xlsx", "wb")
    r = requests.get(k_url)
    xf.write(r.content)
    xf.close()
    ##for garbage collection
    gc.collect()
    return
## Function to create db from the csv files
def db_creation():
    try:
        conn = sqlite3.connect("medicare_hospital_compare.db")
        glob_dir = os.path.join(staging_dir_name, "*.csv")
        for file in glob.glob(glob_dir):
            abs_path = os.path.abspath(file)#gives the absolute path of the file
            try:
                df= pd.read_csv(abs_path)
                column_names = list(df.columns)
                dictionary_type = {y:'str' for y in column_names}
                df = pd.read_csv(abs_path,dtype=dictionary_type)
            except UnicodeDecodeError:
                in_fp = open(abs_path, "rt", encoding='cp1252')  # rt = read text
                input_data = in_fp.read()
                in_fp.close()
                newfile = abs_path + ".fix"
                out_fp = open(newfile, "wt", encoding='utf-8')## conversion  cp1252 encoded file and converts to UTF-8
                for c in input_data:
                    if c != '\0':
                        out_fp.write(c)
                out_fp.close()
                df = pd.read_csv(newfile)
                column_names = list(df.columns)
                dictionary_type = {y:'str' for y in column_names}
                df = pd.read_csv(newfile,dtype=dictionary_type)
            column_names = list(df.columns)
            new_column_names = []
            for each in column_names:
                change = each.lower().replace(" ", "_").replace("-", "_").replace("%", "pct").replace("/", "_")#replacement of the spaces according to db specifications
                if(not change[0].isalpha()):
                    new_column_names.append("c_" + change)
                else:
                    new_column_names.append(change)
            df.columns = new_column_names
            table_name = os.path.splitext(os.path.basename(file))[0].lower().replace(" ","_").replace("-","_").replace("%", "pct").replace("/","_")
            if(not table_name[0].isalpha()):
                table_name = "t_" + table_name
            df.to_sql(table_name,con=conn,if_exists='fail',dtype = {column:'text' for column in df}, index=False)
    ## Raises exception if the connection is made
    except ConnectionError:
        print("Unable to get database connection.. Exiting out of program")
        raise SystemExit
    ## Closes connection upon running the query
    finally:
        conn.close()
    ##Enables garbage collection
    gc.collect()
    return
## function that creates first sheet hospital_ranking.
def output_excel():
    header = ["Provider ID","Hospital Name", "City", "State","County"]
    wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")
    
    in_sheet1 = wb.get_sheet_by_name("Hospital National Ranking")
    in_sheet2 = wb.get_sheet_by_name("Focus States")
    
    first_query = "select provider_id as 'Provider ID', hospital_name as 'Hospital Name'"+\
            ",city as 'City', state as 'State', county_name as 'County' from hospital_general_information where"+ \
            " trim(provider_id) ="
    second_query = "select provider_id as 'Provider ID', hospital_name as 'Hospital Name'" + \
             ",city as 'City', state as 'State', county_name as 'County' from hospital_general_information" +\
              " where state ="
    a =2
    output = []
    conn = sqlite3.connect("medicare_hospital_compare.db")
    while (a<=101):
        qu = first_query + "'"+(in_sheet1.cell(row=a, column=1).value).strip() + "'"
        rows = conn.execute(qu)
        for row in rows:
            output.append(row)
        a += 1
    outputdf = pd.DataFrame(output,columns=header,dtype='str')
    output_sheet =openpyxl.Workbook()
    sheet_1 = output_sheet.create_sheet("Nationwide")
    for r in dataframe_to_rows(outputdf, index=False, header=True):
        sheet_1.append(r)
    b=2
    while in_sheet2.cell(row=b, column=1).value != None:
        states_focused[in_sheet2.cell(row=b, column=1).value.strip()] = in_sheet2.cell(row=b, column=2).value.strip()
        b += 1
    ## For every item in sorted focus states,provider id is checked
    ## looping through result set until count 100 is obtained
    for m , n in sorted(states_focused.items()):
        print("running for ",m)
        sheet_count = 2
        row_count = 1
        res = []
        while (row_count <= 100):
            qu= second_query
            qu = qu + "'"+ n + "' and trim(provider_id)=" + "'" + (in_sheet1.cell(row=sheet_count, column=1).value).strip() + "'"
            rows = conn.execute(qu)
            y = [y for y in rows]
            if(len(y) ==0):
                sheet_count += 1
                continue
            else:
                res.append(y[0])
                sheet_count += 1
                row_count += 1
        outdf_state = pd.DataFrame(res,columns=header,dtype='str')
        sheet_2 = output_sheet.create_sheet(m)
        for r in dataframe_to_rows(outdf_state, index=False, header=True):
            sheet_2.append(r)
    ## first sheet is saved and the database connection is closed
    output_sheet.save("hospital_ranking.xlsx")
    wb = openpyxl.load_workbook("hospital_ranking.xlsx")
    rm = wb.get_sheet_by_name("Sheet")
    wb.remove_sheet(rm)
    print(wb.get_sheet_names())
    wb.save("hospital_ranking.xlsx")
    conn.close()
    ##Enables garbage collection
    gc.collect()
    return
## method for creating measures_statistics. 
def second_sheet():
    result_list= []
    first_query = "select d.measure_id  as 'Measure ID', d.measure_name  as 'Measure Name'," + \
    "min(d.score) as 'Minimum' , max(d.score) as 'Maximum', avg(d.score) as 'Average'," + \
    "AVG((d.score - s.a) * (d.score - s.a)) as 'vari'" + \
    "from timely_and_effective_care___hospital d ," + \
    "(select measure_id as 'm' , avg(score) as 'a' from timely_and_effective_care___hospital group by measure_id) as s " + \
    "WHERE d.measure_id = s.m and " + \
    "CAST(score AS float) IS score group by 1,2;"
    second_query = "select d.measure_id  as 'Measure ID', d.measure_name  as 'Measure Name',"+\
        "min(d.score) as 'Minimum' , max(d.score) as 'Maximum', avg(d.score) as 'Average',"+\
        "AVG((d.score - s.a) * (d.score - s.a)) as 'vari'"+\
        "from timely_and_effective_care___hospital d ,"+\
        "(select measure_id as 'm' , avg(score) as 'a' from timely_and_effective_care___hospital group by measure_id) as s "+\
        "WHERE d.measure_id = s.m and d.state = ? and" +\
        " CAST(score AS float) IS score group by 1,2;"
    conn = sqlite3.connect("medicare_hospital_compare.db")
    headers = ["Measure ID", "Measure Name", "Minimum", "Maximum", "Average", "Standard Deviation"]
    output_sheet = openpyxl.Workbook()
    sheet_1 = output_sheet.create_sheet("Nationwide")
    rows = conn.execute(first_query)
    # appends rows to result
    for row in rows:
        result_list.append(row)
    ## Calculating the standard deviation from variance
    output_dataframe1 = pd.DataFrame(result_list,index=None,columns=headers, dtype='str')
    output_dataframe1['Standard Deviation']= output_dataframe1['Standard Deviation'].astype(float).fillna(0.0) ## Converting the string variance into float
    output_dataframe1['Standard Deviation'] = np.sqrt(output_dataframe1['Standard Deviation'])##calculate sqroot of var using numpy
    output_dataframe1['Standard Deviation'] = output_dataframe1['Standard Deviation'].astype('str') ## Converting into string format for writing it to excel
    for r in dataframe_to_rows(output_dataframe1, index=False, header=True):
        sheet_1.append(r)
    ## query for populating statistical measures for each state
    ##looping over each state and a sheet is created for each state with statistical measures
    for state , stabb in sorted(states_focused.items()):
        result_state = []
        sheet_state = output_sheet.create_sheet(state)
        rows = conn.execute(second_query,(stabb.strip(),))
        for row in rows:
            result_state.append(row)
        ## Calculating the standard deviation from variance
        output_dataframe2 = pd.DataFrame(result_state, index=None, columns=headers, dtype='str') 
        output_dataframe2['Standard Deviation']= output_dataframe2['Standard Deviation'].astype(float).fillna(0.0)## Converting the string variance into float
        output_dataframe2['Standard Deviation'] = np.sqrt(output_dataframe2['Standard Deviation'])##calculate sqroot of var using numpy
        output_dataframe2['Standard Deviation'] = output_dataframe2['Standard Deviation'].astype('str')## Converting into string format for writing it to excel
        for r in dataframe_to_rows(output_dataframe2, index=False, header=True):
            sheet_state.append(r)  
    output_sheet.save("measures_statistics.xlsx")## Save the sheet
    wb = openpyxl.load_workbook("measures_statistics.xlsx")
    rm = wb.get_sheet_by_name("Sheet")
    wb.remove_sheet(rm)
    print(wb.get_sheet_names())
    wb.save("measures_statistics.xlsx")
    conn.close() ## closing the database connection
    ##garbage collection
    gc.collect()
    return
## calling the functions
if __name__ == "__main__":
    csvfiles_download()
    db_creation()
    output_excel()
    second_sheet()
    gc.collect()
